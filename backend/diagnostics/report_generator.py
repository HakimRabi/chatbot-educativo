# Report Generator
# Genera reportes CSV, TXT, JSON y XLSX de los tests de estres

import csv
import io
import json
from datetime import datetime
from typing import Dict, List

# Importar openpyxl para Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportGenerator:
    """Generador de reportes para tests de estres"""
    
    @staticmethod
    def generate_csv(test_data: Dict) -> str:
        """
        Genera un reporte CSV del test
        
        Args:
            test_data: Datos completos del test
        
        Returns:
            String con contenido CSV
        """
        output = io.StringIO()
        
        # Seccion: Informacion General
        output.write("# REPORTE DE TEST DE ESTRES\n")
        output.write(f"# Generado: {datetime.utcnow().isoformat()}Z\n")
        output.write("\n")
        
        # Info del test
        output.write("## INFORMACION DEL TEST\n")
        writer = csv.writer(output)
        writer.writerow(["Campo", "Valor"])
        writer.writerow(["Test ID", test_data.get("test_id", "N/A")])
        writer.writerow(["Nombre", test_data.get("name", "N/A")])
        writer.writerow(["Estado", test_data.get("status", "N/A")])
        writer.writerow(["Duracion (s)", test_data.get("duration_seconds", 0)])
        writer.writerow(["Inicio", test_data.get("started_at", "N/A")])
        writer.writerow(["Fin", test_data.get("completed_at", "N/A")])
        output.write("\n")
        
        # Configuracion
        config = test_data.get("config", {})
        output.write("## CONFIGURACION\n")
        writer.writerow(["Parametro", "Valor"])
        writer.writerow(["Usuarios Concurrentes", config.get("concurrent_users", 0)])
        writer.writerow(["Queries por Usuario", config.get("queries_per_user", 0)])
        writer.writerow(["Duracion Max (s)", config.get("duration_seconds", 0)])
        writer.writerow(["Ramp-up (s)", config.get("ramp_up_seconds", 0)])
        writer.writerow(["Complejidad", config.get("query_complexity", "N/A")])
        writer.writerow(["Modelo", config.get("model_target", "N/A")])
        writer.writerow(["Usar RAG", config.get("use_rag", False)])
        output.write("\n")
        
        # Hardware
        hardware = test_data.get("hardware_info", {})
        output.write("## HARDWARE\n")
        writer.writerow(["Componente", "Valor"])
        writer.writerow(["CPU", hardware.get("cpu_model", "N/A")])
        writer.writerow(["Cores", hardware.get("cpu_cores", 0)])
        writer.writerow(["Threads", hardware.get("cpu_threads", 0)])
        writer.writerow(["GPU", hardware.get("gpu_model", "N/A")])
        writer.writerow(["VRAM (GB)", hardware.get("gpu_vram_gb", 0)])
        writer.writerow(["RAM Total (GB)", hardware.get("ram_total_gb", 0)])
        writer.writerow(["OS", hardware.get("os", "N/A")])
        output.write("\n")
        
        # Resumen
        summary = test_data.get("summary", {})
        output.write("## RESUMEN DE RESULTADOS\n")
        writer.writerow(["Metrica", "Valor"])
        writer.writerow(["Total Queries", summary.get("total_queries", 0)])
        writer.writerow(["Queries Exitosas", summary.get("successful_queries", 0)])
        writer.writerow(["Queries Fallidas", summary.get("failed_queries", 0)])
        writer.writerow(["Tasa de Exito (%)", summary.get("success_rate", 0)])
        output.write("\n")
        
        # Timing
        timing = summary.get("timing", {})
        output.write("## TIEMPOS DE RESPUESTA\n")
        writer.writerow(["Metrica", "Valor (ms)"])
        writer.writerow(["Latencia Promedio", timing.get("avg_latency_ms", 0)])
        writer.writerow(["Latencia Minima", timing.get("min_latency_ms", 0)])
        writer.writerow(["Latencia Maxima", timing.get("max_latency_ms", 0)])
        writer.writerow(["Percentil 50 (Mediana)", timing.get("p50_latency_ms", 0)])
        writer.writerow(["Percentil 95", timing.get("p95_latency_ms", 0)])
        writer.writerow(["Percentil 99", timing.get("p99_latency_ms", 0)])
        output.write("\n")
        
        # Recursos Peak
        peak = summary.get("resources_peak", {})
        output.write("## RECURSOS (PICO MAXIMO)\n")
        writer.writerow(["Recurso", "Valor"])
        writer.writerow(["CPU Max (%)", peak.get("cpu_max_percent", 0)])
        writer.writerow(["RAM Max (%)", peak.get("ram_max_percent", 0)])
        writer.writerow(["RAM Max (MB)", peak.get("ram_max_mb", 0)])
        writer.writerow(["GPU Max (%)", peak.get("gpu_max_percent", 0)])
        writer.writerow(["VRAM Max (MB)", peak.get("vram_max_mb", 0)])
        writer.writerow(["Temperatura Max (C)", peak.get("temperature_max_c", 0)])
        output.write("\n")
        
        # Recursos Average
        avg = summary.get("resources_avg", {})
        output.write("## RECURSOS (PROMEDIO)\n")
        writer.writerow(["Recurso", "Valor"])
        writer.writerow(["CPU Promedio (%)", avg.get("cpu_avg_percent", 0)])
        writer.writerow(["RAM Promedio (%)", avg.get("ram_avg_percent", 0)])
        writer.writerow(["GPU Promedio (%)", avg.get("gpu_avg_percent", 0)])
        writer.writerow(["Temperatura Promedio (C)", avg.get("temperature_avg_c", 0)])
        output.write("\n")
        
        # Throughput
        throughput = summary.get("throughput", {})
        output.write("## RENDIMIENTO\n")
        writer.writerow(["Metrica", "Valor"])
        writer.writerow(["Queries por Segundo", throughput.get("queries_per_second", 0)])
        writer.writerow(["Queries por Minuto", throughput.get("queries_per_minute", 0)])
        output.write("\n")
        
        # Snapshots (serie temporal)
        snapshots = test_data.get("metrics_snapshots", [])
        if snapshots:
            output.write("## SERIE TEMPORAL DE METRICAS\n")
            writer.writerow([
                "Tiempo (s)", "CPU (%)", "RAM (%)", "RAM (MB)", 
                "GPU (%)", "VRAM (MB)", "Temp (C)",
                "Queries Activas", "Queries Completadas", "Queries Fallidas",
                "Latencia Avg (ms)", "Throughput (QPS)"
            ])
            
            for snap in snapshots:
                sys = snap.get("system", {})
                gpu = snap.get("gpu", {})
                perf = snap.get("performance", {})
                
                writer.writerow([
                    snap.get("elapsed_seconds", 0),
                    sys.get("cpu_percent", 0),
                    sys.get("ram_percent", 0),
                    sys.get("ram_used_mb", 0),
                    gpu.get("gpu_percent", 0),
                    gpu.get("vram_used_mb", 0),
                    gpu.get("temperature_c", 0),
                    perf.get("active_queries", 0),
                    perf.get("completed_queries", 0),
                    perf.get("failed_queries", 0),
                    round(perf.get("avg_latency_ms", 0), 2),
                    round(perf.get("throughput_qps", 0), 3)
                ])
        
        return output.getvalue()
    
    @staticmethod
    def generate_txt(test_data: Dict) -> str:
        """
        Genera un reporte TXT legible del test
        
        Args:
            test_data: Datos completos del test
        
        Returns:
            String con contenido TXT
        """
        lines = []
        separator = "=" * 60
        
        lines.append(separator)
        lines.append("REPORTE DE TEST DE ESTRES - CHATBOT EDUCATIVO")
        lines.append(separator)
        lines.append(f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC")
        lines.append("")
        
        # Info del test
        lines.append("-" * 40)
        lines.append("INFORMACION DEL TEST")
        lines.append("-" * 40)
        lines.append(f"  Test ID:  {test_data.get('test_id', 'N/A')}")
        lines.append(f"  Nombre:   {test_data.get('name', 'N/A')}")
        lines.append(f"  Estado:   {test_data.get('status', 'N/A')}")
        lines.append(f"  Duracion: {test_data.get('duration_seconds', 0):.2f} segundos")
        lines.append(f"  Inicio:   {test_data.get('started_at', 'N/A')}")
        lines.append(f"  Fin:      {test_data.get('completed_at', 'N/A')}")
        lines.append("")
        
        # Configuracion
        config = test_data.get("config", {})
        lines.append("-" * 40)
        lines.append("CONFIGURACION DEL TEST")
        lines.append("-" * 40)
        lines.append(f"  Usuarios Concurrentes: {config.get('concurrent_users', 0)}")
        lines.append(f"  Queries por Usuario:   {config.get('queries_per_user', 0)}")
        lines.append(f"  Duracion Maxima:       {config.get('duration_seconds', 0)} s")
        lines.append(f"  Ramp-up:               {config.get('ramp_up_seconds', 0)} s")
        lines.append(f"  Complejidad Query:     {config.get('query_complexity', 'N/A')}")
        lines.append(f"  Modelo Target:         {config.get('model_target', 'N/A')}")
        lines.append(f"  Usar RAG:              {'Si' if config.get('use_rag', False) else 'No'}")
        lines.append("")
        
        # Hardware
        hardware = test_data.get("hardware_info", {})
        lines.append("-" * 40)
        lines.append("HARDWARE DEL SISTEMA")
        lines.append("-" * 40)
        lines.append(f"  CPU:        {hardware.get('cpu_model', 'N/A')}")
        lines.append(f"  Cores:      {hardware.get('cpu_cores', 0)} ({hardware.get('cpu_threads', 0)} threads)")
        lines.append(f"  GPU:        {hardware.get('gpu_model', 'N/A')}")
        lines.append(f"  VRAM:       {hardware.get('gpu_vram_gb', 0)} GB")
        lines.append(f"  RAM Total:  {hardware.get('ram_total_gb', 0)} GB")
        lines.append(f"  OS:         {hardware.get('os', 'N/A')}")
        lines.append("")
        
        # Resumen de resultados
        summary = test_data.get("summary", {})
        lines.append("-" * 40)
        lines.append("RESUMEN DE RESULTADOS")
        lines.append("-" * 40)
        lines.append(f"  Total Queries:     {summary.get('total_queries', 0)}")
        lines.append(f"  Queries Exitosas:  {summary.get('successful_queries', 0)}")
        lines.append(f"  Queries Fallidas:  {summary.get('failed_queries', 0)}")
        lines.append(f"  Tasa de Exito:     {summary.get('success_rate', 0):.1f}%")
        lines.append("")
        
        # Tiempos
        timing = summary.get("timing", {})
        lines.append("-" * 40)
        lines.append("TIEMPOS DE RESPUESTA")
        lines.append("-" * 40)
        lines.append(f"  Latencia Promedio: {timing.get('avg_latency_ms', 0):.2f} ms")
        lines.append(f"  Latencia Minima:   {timing.get('min_latency_ms', 0):.2f} ms")
        lines.append(f"  Latencia Maxima:   {timing.get('max_latency_ms', 0):.2f} ms")
        lines.append(f"  Percentil 50:      {timing.get('p50_latency_ms', 0):.2f} ms")
        lines.append(f"  Percentil 95:      {timing.get('p95_latency_ms', 0):.2f} ms")
        lines.append(f"  Percentil 99:      {timing.get('p99_latency_ms', 0):.2f} ms")
        lines.append("")
        
        # Recursos (pico)
        peak = summary.get("resources_peak", {})
        lines.append("-" * 40)
        lines.append("RECURSOS - PICO MAXIMO")
        lines.append("-" * 40)
        lines.append(f"  CPU Maximo:         {peak.get('cpu_max_percent', 0):.1f}%")
        lines.append(f"  RAM Maximo:         {peak.get('ram_max_percent', 0):.1f}% ({peak.get('ram_max_mb', 0):.0f} MB)")
        lines.append(f"  GPU Maximo:         {peak.get('gpu_max_percent', 0):.1f}%")
        lines.append(f"  VRAM Maximo:        {peak.get('vram_max_mb', 0):.0f} MB")
        lines.append(f"  Temperatura Maxima: {peak.get('temperature_max_c', 0):.0f} C")
        lines.append("")
        
        # Recursos (promedio)
        avg = summary.get("resources_avg", {})
        lines.append("-" * 40)
        lines.append("RECURSOS - PROMEDIO")
        lines.append("-" * 40)
        lines.append(f"  CPU Promedio:         {avg.get('cpu_avg_percent', 0):.1f}%")
        lines.append(f"  RAM Promedio:         {avg.get('ram_avg_percent', 0):.1f}%")
        lines.append(f"  GPU Promedio:         {avg.get('gpu_avg_percent', 0):.1f}%")
        lines.append(f"  Temperatura Promedio: {avg.get('temperature_avg_c', 0):.0f} C")
        lines.append("")
        
        # Throughput
        throughput = summary.get("throughput", {})
        lines.append("-" * 40)
        lines.append("RENDIMIENTO (THROUGHPUT)")
        lines.append("-" * 40)
        lines.append(f"  Queries por Segundo: {throughput.get('queries_per_second', 0):.2f}")
        lines.append(f"  Queries por Minuto:  {throughput.get('queries_per_minute', 0):.2f}")
        lines.append("")
        
        lines.append(separator)
        lines.append("FIN DEL REPORTE")
        lines.append(separator)
        
        return "\n".join(lines)
    
    @staticmethod
    def generate_json(test_data: Dict) -> str:
        """
        Genera un reporte JSON del test
        
        Args:
            test_data: Datos completos del test
        
        Returns:
            String con contenido JSON
        """
        return json.dumps(test_data, indent=2, default=str)
    
    @staticmethod
    def generate_xlsx(test_data: Dict) -> bytes:
        """
        Genera un reporte Excel (.xlsx) con multiples hojas bien formateadas
        
        Args:
            test_data: Datos completos del test
        
        Returns:
            Bytes del archivo Excel
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no esta instalado. Ejecutar: pip install openpyxl")
        
        wb = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        section_font = Font(bold=True, size=12)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        number_alignment = Alignment(horizontal="right")
        
        def style_header_row(ws, row, col_count):
            """Aplica estilo a fila de encabezado"""
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
        
        def auto_width(ws):
            """Ajusta ancho de columnas automaticamente"""
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        if cell.value and not isinstance(cell, type(None)):
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = max(adjusted_width, 12)
        
        # ========== HOJA 1: RESUMEN ==========
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        
        row = 1
        
        # Titulo
        ws_resumen.cell(row=row, column=1, value="REPORTE DE TEST DE ESTRES")
        ws_resumen.cell(row=row, column=1).font = Font(bold=True, size=16)
        ws_resumen.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
        
        ws_resumen.cell(row=row, column=1, value=f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC")
        row += 2
        
        # Informacion del Test
        ws_resumen.cell(row=row, column=1, value="INFORMACION DEL TEST")
        ws_resumen.cell(row=row, column=1).font = section_font
        ws_resumen.cell(row=row, column=1).fill = section_fill
        ws_resumen.cell(row=row, column=2).fill = section_fill
        row += 1
        
        info_data = [
            ("Test ID", test_data.get("test_id", "N/A")),
            ("Nombre", test_data.get("name", "N/A")),
            ("Estado", test_data.get("status", "N/A")),
            ("Duracion (segundos)", test_data.get("duration_seconds", 0)),
            ("Inicio", str(test_data.get("started_at", "N/A"))),
            ("Fin", str(test_data.get("completed_at", "N/A"))),
        ]
        
        for label, value in info_data:
            ws_resumen.cell(row=row, column=1, value=label).border = thin_border
            ws_resumen.cell(row=row, column=2, value=value).border = thin_border
            row += 1
        
        row += 1
        
        # Configuracion
        config = test_data.get("config", {})
        ws_resumen.cell(row=row, column=1, value="CONFIGURACION")
        ws_resumen.cell(row=row, column=1).font = section_font
        ws_resumen.cell(row=row, column=1).fill = section_fill
        ws_resumen.cell(row=row, column=2).fill = section_fill
        row += 1
        
        config_data = [
            ("Usuarios Concurrentes", config.get("concurrent_users", 0)),
            ("Queries por Usuario", config.get("queries_per_user", 0)),
            ("Duracion Maxima (s)", config.get("duration_seconds", 0)),
            ("Ramp-up (s)", config.get("ramp_up_seconds", 0)),
            ("Complejidad Query", config.get("query_complexity", "N/A")),
            ("Modelo Target", config.get("model_target", "N/A")),
            ("Usar RAG", "Si" if config.get("use_rag", False) else "No"),
        ]
        
        for label, value in config_data:
            ws_resumen.cell(row=row, column=1, value=label).border = thin_border
            ws_resumen.cell(row=row, column=2, value=value).border = thin_border
            row += 1
        
        row += 1
        
        # Hardware
        hardware = test_data.get("hardware_info", {})
        ws_resumen.cell(row=row, column=1, value="HARDWARE")
        ws_resumen.cell(row=row, column=1).font = section_font
        ws_resumen.cell(row=row, column=1).fill = section_fill
        ws_resumen.cell(row=row, column=2).fill = section_fill
        row += 1
        
        hardware_data = [
            ("CPU", hardware.get("cpu_model", "N/A")),
            ("Cores", hardware.get("cpu_cores", 0)),
            ("Threads", hardware.get("cpu_threads", 0)),
            ("GPU", hardware.get("gpu_model", "N/A")),
            ("VRAM (GB)", hardware.get("gpu_vram_gb", 0)),
            ("RAM Total (GB)", hardware.get("ram_total_gb", 0)),
            ("Sistema Operativo", hardware.get("os", "N/A")),
        ]
        
        for label, value in hardware_data:
            ws_resumen.cell(row=row, column=1, value=label).border = thin_border
            ws_resumen.cell(row=row, column=2, value=value).border = thin_border
            row += 1
        
        auto_width(ws_resumen)
        
        # ========== HOJA 2: RESULTADOS ==========
        ws_resultados = wb.create_sheet("Resultados")
        
        row = 1
        summary = test_data.get("summary", {})
        
        # Resumen Queries
        ws_resultados.cell(row=row, column=1, value="RESUMEN DE QUERIES")
        ws_resultados.cell(row=row, column=1).font = section_font
        ws_resultados.cell(row=row, column=1).fill = section_fill
        ws_resultados.cell(row=row, column=2).fill = section_fill
        row += 1
        
        queries_data = [
            ("Total Queries", summary.get("total_queries", 0)),
            ("Queries Exitosas", summary.get("successful_queries", 0)),
            ("Queries Fallidas", summary.get("failed_queries", 0)),
            ("Tasa de Exito (%)", summary.get("success_rate", 0)),
        ]
        
        for label, value in queries_data:
            ws_resultados.cell(row=row, column=1, value=label).border = thin_border
            ws_resultados.cell(row=row, column=2, value=value).border = thin_border
            row += 1
        
        row += 1
        
        # Tiempos de Respuesta
        timing = summary.get("timing", {})
        ws_resultados.cell(row=row, column=1, value="TIEMPOS DE RESPUESTA (ms)")
        ws_resultados.cell(row=row, column=1).font = section_font
        ws_resultados.cell(row=row, column=1).fill = section_fill
        ws_resultados.cell(row=row, column=2).fill = section_fill
        row += 1
        
        timing_data = [
            ("Latencia Promedio", timing.get("avg_latency_ms", 0)),
            ("Latencia Minima", timing.get("min_latency_ms", 0)),
            ("Latencia Maxima", timing.get("max_latency_ms", 0)),
            ("Percentil 50 (Mediana)", timing.get("p50_latency_ms", 0)),
            ("Percentil 95", timing.get("p95_latency_ms", 0)),
            ("Percentil 99", timing.get("p99_latency_ms", 0)),
        ]
        
        for label, value in timing_data:
            ws_resultados.cell(row=row, column=1, value=label).border = thin_border
            ws_resultados.cell(row=row, column=2, value=value).border = thin_border
            row += 1
        
        row += 1
        
        # Recursos Pico
        peak = summary.get("resources_peak", {})
        ws_resultados.cell(row=row, column=1, value="RECURSOS - PICO MAXIMO")
        ws_resultados.cell(row=row, column=1).font = section_font
        ws_resultados.cell(row=row, column=1).fill = section_fill
        ws_resultados.cell(row=row, column=2).fill = section_fill
        row += 1
        
        peak_data = [
            ("CPU Maximo (%)", peak.get("cpu_max_percent", 0)),
            ("RAM Maximo (%)", peak.get("ram_max_percent", 0)),
            ("RAM Maximo (MB)", peak.get("ram_max_mb", 0)),
            ("GPU Maximo (%)", peak.get("gpu_max_percent", 0)),
            ("VRAM Maximo (MB)", peak.get("vram_max_mb", 0)),
            ("Temperatura Maxima (C)", peak.get("temperature_max_c", 0)),
        ]
        
        for label, value in peak_data:
            ws_resultados.cell(row=row, column=1, value=label).border = thin_border
            ws_resultados.cell(row=row, column=2, value=value).border = thin_border
            row += 1
        
        row += 1
        
        # Recursos Promedio
        avg = summary.get("resources_avg", {})
        ws_resultados.cell(row=row, column=1, value="RECURSOS - PROMEDIO")
        ws_resultados.cell(row=row, column=1).font = section_font
        ws_resultados.cell(row=row, column=1).fill = section_fill
        ws_resultados.cell(row=row, column=2).fill = section_fill
        row += 1
        
        avg_data = [
            ("CPU Promedio (%)", avg.get("cpu_avg_percent", 0)),
            ("RAM Promedio (%)", avg.get("ram_avg_percent", 0)),
            ("GPU Promedio (%)", avg.get("gpu_avg_percent", 0)),
            ("Temperatura Promedio (C)", avg.get("temperature_avg_c", 0)),
        ]
        
        for label, value in avg_data:
            ws_resultados.cell(row=row, column=1, value=label).border = thin_border
            ws_resultados.cell(row=row, column=2, value=value).border = thin_border
            row += 1
        
        row += 1
        
        # Throughput
        throughput = summary.get("throughput", {})
        ws_resultados.cell(row=row, column=1, value="RENDIMIENTO")
        ws_resultados.cell(row=row, column=1).font = section_font
        ws_resultados.cell(row=row, column=1).fill = section_fill
        ws_resultados.cell(row=row, column=2).fill = section_fill
        row += 1
        
        throughput_data = [
            ("Queries por Segundo", throughput.get("queries_per_second", 0)),
            ("Queries por Minuto", throughput.get("queries_per_minute", 0)),
        ]
        
        for label, value in throughput_data:
            ws_resultados.cell(row=row, column=1, value=label).border = thin_border
            ws_resultados.cell(row=row, column=2, value=value).border = thin_border
            row += 1
        
        auto_width(ws_resultados)
        
        # ========== HOJA 3: SERIE TEMPORAL ==========
        snapshots = test_data.get("metrics_snapshots", [])
        if snapshots:
            ws_temporal = wb.create_sheet("Serie Temporal")
            
            # Headers
            headers = [
                "Tiempo (s)", "CPU (%)", "RAM (%)", "RAM (MB)",
                "GPU (%)", "VRAM (MB)", "Temperatura (C)",
                "Queries Activas", "Queries Completadas", "Queries Fallidas",
                "Latencia Avg (ms)", "Throughput (QPS)"
            ]
            
            for col, header in enumerate(headers, 1):
                ws_temporal.cell(row=1, column=col, value=header)
            
            style_header_row(ws_temporal, 1, len(headers))
            
            # Datos
            for row_idx, snap in enumerate(snapshots, 2):
                sys_data = snap.get("system", {})
                gpu_data = snap.get("gpu", {})
                perf_data = snap.get("performance", {})
                
                values = [
                    snap.get("elapsed_seconds", 0),
                    sys_data.get("cpu_percent", 0),
                    sys_data.get("ram_percent", 0),
                    sys_data.get("ram_used_mb", 0),
                    gpu_data.get("gpu_percent", 0),
                    gpu_data.get("vram_used_mb", 0),
                    gpu_data.get("temperature_c", 0),
                    perf_data.get("active_queries", 0),
                    perf_data.get("completed_queries", 0),
                    perf_data.get("failed_queries", 0),
                    round(perf_data.get("avg_latency_ms", 0), 2),
                    round(perf_data.get("throughput_qps", 0), 4)
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws_temporal.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border
                    if isinstance(value, (int, float)):
                        cell.alignment = number_alignment
            
            auto_width(ws_temporal)
        
        # Guardar a bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
